import cv2
from pyzbar import pyzbar
import openpyxl


def ReadQR():
    # Open the camera
    cap = cv2.VideoCapture(0)

    # Load the workbook or create a new one if it doesn't exist
    try:
        wb = openpyxl.load_workbook("qr_data.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Select the active worksheet
    ws = wb.active

    # Get the current maximum row number in the worksheet
    max_row = ws.max_row

    while True:
        # Read the frame from the camera
        _, frame = cap.read()

        # Decode the QR code in the frame
        decoded_objs = pyzbar.decode(frame)

        # Display the frame
        cv2.imshow("Frame", frame)

        # If a QR code is detected, print the data
        for obj in decoded_objs:
            data = obj.data.decode()
            print("Data:", data)

            # Split the data into a list of strings using '#'
            strings = data.split('#')

            # Write the strings to the Excel sheet
            for i, string in enumerate(strings):
                row = max_row + 1
                ws.cell(row=row, column=i+1, value=string)

            # Increment the row number for the next iteration
            max_row += 1

            # Save the workbook and exit the loop
            wb.save("qr_data.xlsx")
            cap.release()
            cv2.destroyAllWindows()
            return

        # Wait for key press
        key = cv2.waitKey(1) & 0xFF

        # If 'q' is pressed, exit the loop
        if key == ord('q'):
            break

    # Release the camera and close the window
    cap.release()


ReadQR()